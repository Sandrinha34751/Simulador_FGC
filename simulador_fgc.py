# --- Importação de bibliotecas necessárias ---
import tkinter as tk  # Interface gráfica
from tkinter import ttk, messagebox  # Widgets e mensagens
import matplotlib.pyplot as plt  # Gráficos
import pandas as pd  # Manipulação de dados
import os  # Operações com arquivos
from openpyxl import load_workbook  # Leitura/edição de arquivos Excel
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side  # Estilos do Excel

# --- Constante do limite de cobertura do FGC ---
FGC_LIMITE = 250000  # Valor máximo garantido por CPF/CNPJ por instituição

# --- Função que calcula o valor garantido pelo FGC ---
def calcular_garantia(valor_investido):
    return min(valor_investido, FGC_LIMITE)

# --- Função que exibe o gráfico de barras com os valores garantido e não garantido ---
def exibir_grafico(valor_investido, valor_garantido):
    valor_nao_garantido = max(valor_investido - valor_garantido, 0)
    categorias = ['Garantido', 'Não Garantido']
    valores = [valor_garantido, valor_nao_garantido]
    cores = ['#4CAF50', '#F44336']

    # Criando o gráfico
    plt.figure(figsize=(6, 4))
    barras = plt.bar(categorias, valores, color=cores)

    # Adicionando os valores acima das barras
    for barra in barras:
        altura = barra.get_height()
        plt.annotate(f'R$ {altura:,.2f}', 
                     xy=(barra.get_x() + barra.get_width() / 2, altura),
                     xytext=(0, 3), 
                     textcoords="offset points",
                     ha='center', va='bottom')

    plt.title('Garantia FGC')
    plt.ylabel('Valor (R$)')
    plt.ylim(0, max(valor_investido * 1.1, 1000))
    plt.tight_layout()
    plt.show()

# --- Classe principal que define a interface gráfica da aplicação ---
class SimuladorFGCApp:
    def __init__(self, root):
        # Configurações da janela principal
        self.root = root
        self.root.title("Simulador de FGC")
        self.root.geometry("520x570")
        self.root.configure(bg="#e6f2ff")

        # Estilização dos componentes
        estilo = ttk.Style()
        estilo.configure("TLabel", background="#e6f2ff", font=("Arial", 12))
        estilo.configure("TButton", font=("Arial", 12, "bold"), padding=6)
        estilo.configure("TEntry", padding=5)
        estilo.configure("TCombobox", padding=5)

        # Título da aplicação
        ttk.Label(root, text="💰 Simulador de FGC 💰", font=("Arial", 20, "bold")).pack(pady=20)

        # Container com os campos de entrada
        container = ttk.Frame(root, padding=10)
        container.pack(pady=10)

        # Campo para valor investido
        ttk.Label(container, text="➡️ Valor do investimento (R$):").grid(row=0, column=0, sticky='w', pady=5)
        self.valor_entry = ttk.Entry(container, width=30)
        self.valor_entry.grid(row=0, column=1, pady=5)

        # ComboBox para tipo de instituição
        ttk.Label(container, text="🏦 Tipo de instituição:").grid(row=1, column=0, sticky='w', pady=5)
        self.instituicao_var = tk.StringVar()
        self.instituicao_combo = ttk.Combobox(container, textvariable=self.instituicao_var, state="readonly", width=28)
        self.instituicao_combo['values'] = ("Banco", "Financeira", "Cooperativa")
        self.instituicao_combo.current(0)
        self.instituicao_combo.grid(row=1, column=1, pady=5)

        # ComboBox para tipo de aplicação
        ttk.Label(container, text="📄 Tipo de aplicação:").grid(row=2, column=0, sticky='w', pady=5)
        self.aplicacao_var = tk.StringVar()
        self.aplicacao_combo = ttk.Combobox(container, textvariable=self.aplicacao_var, state="readonly", width=28)
        self.aplicacao_combo['values'] = ("CDB", "LCI", "LCA", "Poupança", "Outros")
        self.aplicacao_combo.current(0)
        self.aplicacao_combo.grid(row=2, column=1, pady=5)

        # Botão para simular a garantia
        self.simular_btn = ttk.Button(root, text="✅ Simular Garantia", command=self.simular)
        self.simular_btn.pack(pady=15)

        # Label para exibir o resultado da simulação
        self.resultado_label = ttk.Label(root, text="", font=("Arial", 12), background="#d9edf7", wraplength=450, padding=10)
        self.resultado_label.pack(pady=10, fill='x', padx=20)

        # Botão para mostrar o gráfico
        self.grafico_btn = ttk.Button(root, text="📊 Mostrar Gráfico", command=self.mostrar_grafico)
        self.grafico_btn.pack(pady=10)
        self.grafico_btn['state'] = 'disabled'  # Desabilitado até que uma simulação seja feita

        # Botão para exportar os dados para Excel
        self.excel_btn = ttk.Button(root, text="📤 Exportar para Excel", command=self.exportar_excel)
        self.excel_btn.pack(pady=5)
        self.excel_btn['state'] = 'disabled'

        # Botão com informações sobre o FGC
        self.sobre_btn = ttk.Button(root, text="ℹ️ Sobre o FGC", command=self.mostrar_info)
        self.sobre_btn.pack(pady=5)

        # Variáveis para armazenar os valores da simulação
        self.valor_investido = 0
        self.valor_garantido = 0
        self.valor_nao_garantido = 0

    # --- Função que realiza a simulação ---
    def simular(self):
        try:
            valor = float(self.valor_entry.get())
            if valor <= 0:
                raise ValueError("Valor deve ser maior que zero.")

            # Calcula os valores garantido e não garantido
            garantia = calcular_garantia(valor)
            nao_garantido = max(valor - garantia, 0)

            # Exibe os resultados na interface
            mensagem = (
                f"⭐ Tipo: {self.aplicacao_var.get()} | Instituição: {self.instituicao_var.get()}\n"
                f"⭐ Valor investido: R$ {valor:,.2f}\n"
                f"⭐ Valor garantido pelo FGC: R$ {garantia:,.2f}\n"
                f"⭐ Valor NÃO garantido: R$ {nao_garantido:,.2f}"
            )
            self.resultado_label.config(text=mensagem)

            # Atualiza variáveis e ativa botões
            self.valor_investido = valor
            self.valor_garantido = garantia
            self.valor_nao_garantido = nao_garantido
            self.grafico_btn['state'] = 'normal'
            self.excel_btn['state'] = 'normal'

        except ValueError as e:
            messagebox.showerror("Erro", str(e))

    # --- Exibe o gráfico de barras ---
    def mostrar_grafico(self):
        exibir_grafico(self.valor_investido, self.valor_garantido)

    # --- Exporta os dados da simulação para um arquivo Excel ---
    def exportar_excel(self):
        nome_arquivo = "relatorio_fgc.xlsx"

        nova_linha = {
            "Tipo de Aplicação": self.aplicacao_var.get(),
            "Instituição": self.instituicao_var.get(),
            "Valor Investido (R$)": self.valor_investido,
            "Valor Garantido FGC (R$)": self.valor_garantido,
            "Valor Não Garantido (R$)": self.valor_nao_garantido
        }

        # Se o arquivo já existe, adiciona nova linha
        if os.path.exists(nome_arquivo):
            df_existente = pd.read_excel(nome_arquivo)
            df_novo = pd.concat([df_existente, pd.DataFrame([nova_linha])], ignore_index=True)
        else:
            df_novo = pd.DataFrame([nova_linha])

        # Salva no Excel
        df_novo.to_excel(nome_arquivo, index=False)

        # Ajusta o estilo da planilha
        wb = load_workbook(nome_arquivo)
        ws = wb.active

        header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Estiliza os cabeçalhos
        for i, cell in enumerate(ws[1], start=1):
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Formata os valores em reais e adiciona bordas
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=5):
            for cell in row:
                cell.number_format = 'R$ #,##0.00'
                cell.border = border
                cell.alignment = Alignment(horizontal="right")

        # Salva o arquivo final
        wb.save(nome_arquivo)
        messagebox.showinfo("Exportação", f"Simulação salva em '{nome_arquivo}' com sucesso!")

    # --- Exibe informações sobre o FGC ---
    def mostrar_info(self):
        texto = (
         "🔎 O que é o FGC?\n"
            "O Fundo Garantidor de Créditos (FGC) é uma entidade que protege parte do dinheiro investido por pessoas físicas em instituições financeiras.\n\n"
            "👥 Quem tem direito?\n"
            "- Pessoas físicas e jurídicas com investimentos em produtos cobertos.\n\n"
            "💼 Aplicações cobertas:\n"
            "- CDB, LCI, LCA, Contas Correntes, entre outros.\n\n"
            "💸 Limite de cobertura:\n"
            "- Até R$ 250.000 por CPF/CNPJ por instituição financeira."
        )
        messagebox.showinfo("Sobre o FGC", texto)

# --- Execução da aplicação ---
if __name__ == "__main__":
    root = tk.Tk()
    app = SimuladorFGCApp(root)
    root.mainloop()
